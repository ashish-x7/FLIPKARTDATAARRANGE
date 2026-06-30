import os
import io
import pandas as pd
import zipfile
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
CORS(app)

# Detect file type by inspecting magic bytes
def detect_file_type(file_bytes, filename):
    if len(file_bytes) > 4:
        # ZIP signature (xlsx)
        if file_bytes[:4] == b'PK\x03\x04':
            return 'xlsx'
        # OLECF signature (xls binary format - old Excel)
        if file_bytes[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1':
            return 'xls'
    # Fallback to extension check
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    if ext in ['xlsx', 'xls', 'csv']:
        return ext
    return 'csv'

# Check if file bytes represent an HTML spreadsheet
def is_html_bytes(file_bytes):
    encodings = ['utf-8', 'utf-16', 'latin-1', 'cp1252']
    for enc in encodings:
        try:
            sample = file_bytes[:1000].decode(enc, errors='ignore').strip().lower()
            # Strip potential BOM characters
            sample = sample.lstrip('\ufeff\xff\xfe')
            # Common HTML tags or spreadsheet namespaces
            if (sample.startswith('<html') or 
                sample.startswith('<!doc') or 
                sample.startswith('<table') or 
                sample.startswith('<tr') or
                sample.startswith('<head') or
                sample.startswith('<xml') or
                '<table' in sample or
                '<html' in sample or
                'xmlns:x="urn:schemas-microsoft-com:office:excel"' in sample):
                return True
        except Exception:
            continue
    return False

# Parse HTML Excel sheet using pandas read_html
def parse_html_spreadsheet(file_bytes):
    dfs = pd.read_html(io.BytesIO(file_bytes))
    if dfs:
        df = dfs[0]
        return df.astype(str)
    raise ValueError("No table found in HTML spreadsheet")

# Robust CSV Reader that auto-detects encoding and separator
def read_csv_robustly(file_bytes, header='infer'):
    encodings = ['utf-8', 'utf-16', 'latin-1', 'cp1252']
    decoded_str = None
    best_enc = 'utf-8'
    
    # Try different encodings to decode the byte contents
    for enc in encodings:
        try:
            decoded_str = file_bytes.decode(enc)
            best_enc = enc
            break
        except Exception:
            continue
            
    if not decoded_str:
        # Fallback to default pandas reader
        return pd.read_csv(io.BytesIO(file_bytes), header=header, dtype=str)
        
    # Detect separator based on character counts in the first few lines
    lines = [line for line in decoded_str.split('\n') if line.strip()][:5]
    sample = "\n".join(lines)
    
    delimiters = [',', '\t', ';', '|']
    best_delim = ','
    max_count = 0
    for d in delimiters:
        count = sample.count(d)
        if count > max_count:
            max_count = count
            best_delim = d
            
    try:
        df = pd.read_csv(
            io.BytesIO(file_bytes),
            sep=best_delim,
            header=header,
            encoding=best_enc,
            dtype=str
        )
        return df
    except Exception:
        pass
        
    return pd.read_csv(io.BytesIO(file_bytes), header=header, dtype=str)

# Excel COM-based repair/conversion tool (utilizes local Excel to repair malformed sheets)
def convert_xls_to_xlsx_via_excel(file_bytes, filename):
    try:
        import pythoncom
        import win32com.client
    except (ImportError, ModuleNotFoundError) as e:
        print(f"[REPAIR FAIL] Excel COM conversion is not available on this platform: {e}", flush=True)
        return None
    
    os.makedirs('temp', exist_ok=True)
    temp_xls = os.path.join('temp', 'temp_input.xls')
    temp_xlsx = os.path.join('temp', 'temp_output.xlsx')
    
    # Clean previous temp files
    for path in [temp_xls, temp_xlsx]:
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
                
    # Write bytes to temp input file
    with open(temp_xls, 'wb') as f:
        f.write(file_bytes)
        
    excel = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        abs_xls = os.path.abspath(temp_xls)
        abs_xlsx = os.path.abspath(temp_xlsx)
        
        wb = excel.Workbooks.Open(abs_xls)
        # FileFormat=51 is standard XLSX
        wb.SaveAs(abs_xlsx, FileFormat=51)
        wb.Close(SaveChanges=False)
        excel.Quit()
        
        # Read the repaired XLSX file bytes
        with open(temp_xlsx, 'rb') as f:
            xlsx_bytes = f.read()
            
        # Clean up
        for path in [temp_xls, temp_xlsx]:
            try:
                os.remove(path)
            except Exception:
                pass
                
        return xlsx_bytes
    except Exception as e:
        print(f"[REPAIR FAIL] Excel COM conversion failed: {e}", flush=True)
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        return None

# Detect sheet name matching 'order' or 'orders' case-insensitively, falling back to the first sheet
def read_excel_by_detecting_sheet(file_bytes, engine, header):
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine=engine)
    sheet_name = None
    
    # Clean and check sheet names for "order" or "orders"
    for s in xls.sheet_names:
        s_clean = s.strip().lower()
        if s_clean == 'order' or s_clean == 'orders':
            sheet_name = s
            break
            
    # Secondary check: search for "order" substring inside sheet names
    if not sheet_name:
        for s in xls.sheet_names:
            if 'order' in s.strip().lower():
                sheet_name = s
                break
                
    # Fallback to the first sheet in the list
    if not sheet_name:
        sheet_name = xls.sheet_names[0]
        
    print(f"[DEBUG Excel] Parsing sheet: '{sheet_name}' from workbook sheets: {xls.sheet_names}", flush=True)
    return xls.parse(sheet_name, header=header, dtype=str)

# Unified reader function that handles XLSX, XLS, HTML-spreadsheets, and CSV/TSV
def load_any_sheet_to_dataframe(file_bytes, filename, header=None):
    filetype = detect_file_type(file_bytes, filename)
    
    if is_html_bytes(file_bytes):
        filetype = 'html'
        
    print(f"[DEBUG Loader] Filename: {filename}, Detected Format: {filetype}", flush=True)
    
    if filetype == 'xlsx':
        try:
            return read_excel_by_detecting_sheet(file_bytes, engine='openpyxl', header=header)
        except Exception as e:
            print(f"XLSX load fail on {filename}: {e}. Trying CSV fallback...", flush=True)
            
    if filetype == 'xls':
        # Try standard xlrd engine first
        try:
            return read_excel_by_detecting_sheet(file_bytes, engine='xlrd', header=header)
        except Exception as e:
            print(f"XLS (xlrd) load fail on {filename}: {e}. Attempting Excel COM repair...", flush=True)
            xlsx_bytes = convert_xls_to_xlsx_via_excel(file_bytes, filename)
            if xlsx_bytes:
                try:
                    print(f"COM Repair succeeded for {filename}. Loading XLSX...", flush=True)
                    return read_excel_by_detecting_sheet(xlsx_bytes, engine='openpyxl', header=header)
                except Exception as ex:
                    print(f"Failed to load repaired XLSX for {filename}: {ex}", flush=True)
            else:
                print(f"Excel COM repair returned None for {filename}", flush=True)
            
    if filetype == 'html':
        try:
            df = parse_html_spreadsheet(file_bytes)
            # Standardize headers if header=0 is specified
            if header == 0 and len(df) > 0:
                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)
            return df
        except Exception as e:
            print(f"HTML load fail on {filename}: {e}. Trying CSV fallback...", flush=True)
            
    # Fallback to robust CSV
    try:
        return read_csv_robustly(file_bytes, header=header)
    except Exception as e:
        print(f"CSV load fail on {filename}: {e}", flush=True)
        # Try COM repair as a last resort
        print(f"Last resort: Attempting Excel COM repair for {filename}...", flush=True)
        xlsx_bytes = convert_xls_to_xlsx_via_excel(file_bytes, filename)
        if xlsx_bytes:
            try:
                return read_excel_by_detecting_sheet(xlsx_bytes, engine='openpyxl', header=header)
            except Exception as ex:
                print(f"Last resort load failed: {ex}", flush=True)
        raise e

# Helper function to find a column by name or index
def find_col_key(df, target_names, default_idx):
    cols = list(df.columns)
    for col in cols:
        if str(col).strip().lower() in target_names:
            return col
    # Fallback to column index if within range
    if default_idx < len(cols):
        return cols[default_idx]
    return None

# Cleaning functions for Merging Tab (Supports CSV & Excel)
def clean_order_item_id(val):
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val).strip()
    # Remove "OI:" or "oi:" prefix case-insensitively
    if val_str.upper().startswith("OI:"):
        val_str = val_str[3:]
    return val_str.strip()

def clean_quotes(val):
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val).strip()
    # Remove all double quotes
    return val_str.replace('"', '').strip()

def clean_sku(val):
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val).strip()
    # Remove all double quotes
    val_str = val_str.replace('"', '').strip()
    # Remove "SKU:" prefix case-insensitively if it exists
    if val_str.upper().startswith("SKU:"):
        val_str = val_str[4:]
    return val_str.strip()

# Helpers for Renaming Tab (Fuzzy check and CSV read compatibility)
def clean_string_for_matching(s):
    if pd.isna(s) or s is None:
        return ""
    val_str = str(s).upper()
    for char in [' ', '-', '_', '.']:
        val_str = val_str.replace(char, '')
    return val_str.strip()

def is_mapping_file(filename, file_bytes):
    fname_lower = filename.lower()
    # 1. Filename heuristic
    if 'arrange' in fname_lower or 'mapping' in fname_lower:
        return True
        
    # 2. Check if file contains a sheet named ARRANGE
    filetype = detect_file_type(file_bytes, filename)
    if is_html_bytes(file_bytes):
        filetype = 'html'
        
    if filetype in ['xlsx', 'xls']:
        try:
            xls = pd.ExcelFile(io.BytesIO(file_bytes))
            for s in xls.sheet_names:
                if s.strip().upper() == 'ARRANGE':
                    return True
        except Exception:
            pass
    return False

def parse_mapping_df(file_bytes, filename):
    try:
        df = load_any_sheet_to_dataframe(file_bytes, filename, header=None)
        if df is not None:
            df = df.dropna(subset=[0, 1])
        return df
    except Exception as e:
        print(f"Error parsing mapping file {filename}: {e}", flush=True)
        return None

def load_saved_mapping():
    os.makedirs('temp', exist_ok=True)
    for ext in ['.xlsx', '.xls', '.csv']:
        path = os.path.join('temp', f'arrange_mapping{ext}')
        if os.path.exists(path):
            try:
                with open(path, 'rb') as f:
                    file_bytes = f.read()
                return parse_mapping_df(file_bytes, f'arrange_mapping{ext}')
            except Exception:
                pass
    return None

def find_rename_code_option_a_from_bytes(file_bytes, filename, mapping_df):
    try:
        # Load headerless table structure
        df = load_any_sheet_to_dataframe(file_bytes, filename, header=None)
        
        p2_val = None
        # Cell P2 is row index 1, column index 15
        if len(df) > 1 and len(df.columns) > 15:
            p2_val = df.iloc[1, 15]
            
        if p2_val is None:
            return ""
            
        rename_code = str(p2_val).strip()
        if not rename_code:
            return ""
            
        if mapping_df is None or len(mapping_df) == 0:
            return rename_code
            
        find_name = clean_string_for_matching(rename_code)
        if not find_name:
            return rename_code
            
        # Search in mapping (index 0 is code, index 1 is brand name)
        for idx, row in mapping_df.iterrows():
            if len(row) >= 2:
                code_val = str(row.iloc[0]).strip()
                name_val = str(row.iloc[1]).strip()
                
                arr_name = clean_string_for_matching(name_val)
                if not arr_name:
                    continue
                    
                if (arr_name in find_name) or (find_name in arr_name):
                    return code_val
                    
        return rename_code
    except Exception as e:
        print(f"Error extracting Option A from {filename}: {e}", flush=True)
        return ""

def find_rename_code_option_b_from_bytes(file_bytes, filename):
    try:
        # Load headerless table structure
        df = load_any_sheet_to_dataframe(file_bytes, filename, header=None)
        print(f"[DEBUG Option B] Shape of loaded spreadsheet: {df.shape}", flush=True)
        
        # Find column index of invoice number (default: Column G = index 6)
        col_idx = 6
        if len(df) > 0:
            first_row = [str(x).strip().lower() for x in df.iloc[0]]
            for pos, val in enumerate(first_row):
                if 'invoice' in val:
                    col_idx = pos
                    print(f"[DEBUG Option B] Detected invoice column at index {col_idx} based on header '{val}'", flush=True)
                    break
        
        if len(df.columns) > col_idx:
            for idx in range(1, len(df)):
                cell_val = df.iloc[idx, col_idx]
                if pd.notna(cell_val):
                    val_str = str(cell_val).strip()
                    if val_str and not val_str.upper().startswith("CGJ1-"):
                        first_part = val_str.split('-')[0].strip()
                        return first_part[-3:]
    except Exception as e:
        print(f"Error reading Column G from {filename}: {e}", flush=True)
        
    return ""


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_files():
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files[]')
    if not files or len(files) == 0 or files[0].filename == '':
        return jsonify({'error': 'No selected files'}), 400

    dataframes = []
    
    for file in files:
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls') or file.filename.endswith('.csv')):
            continue
        
        try:
            file_bytes = file.read()
            # Merged tab requires headers
            df = load_any_sheet_to_dataframe(file_bytes, file.filename, header=0)
            
            if df.empty:
                continue
                
            # Clean columns
            col_id = find_col_key(df, ['order_item_id', 'order item id', 'orderitemid'], 0)
            if col_id is not None:
                df[col_id] = df[col_id].apply(clean_order_item_id)
                
            col_sku = find_col_key(df, ['sku', 'seller sku', 'sellersku'], 7)
            if col_sku is not None:
                df[col_sku] = df[col_sku].apply(clean_sku)
                
            col_title = find_col_key(df, ['product_title', 'product title', 'producttitle', 'title'], 9)
            if col_title is not None:
                df[col_title] = df[col_title].apply(clean_quotes)
            
            dataframes.append(df)
            
        except Exception as e:
            return jsonify({'error': f'Error reading {file.filename}: {str(e)}'}), 500

    if not dataframes:
        return jsonify({'error': 'No valid order sheets found in uploaded files.'}), 400

    try:
        combined_df = pd.concat(dataframes, ignore_index=True)
    except Exception as e:
        return jsonify({'error': f'Error merging data sheets: {str(e)}'}), 500

    combined_df = combined_df.fillna('')
    preview_data = combined_df.head(10).to_dict(orient='records')
    columns = list(combined_df.columns)
    
    os.makedirs('temp', exist_ok=True)
    temp_path = os.path.join('temp', 'merged_output.xlsx')
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        order_item_col_idx = -1
        col_id_name = find_col_key(combined_df, ['order_item_id', 'order item id', 'orderitemid'], 0)
        if col_id_name in combined_df.columns:
            order_item_col_idx = list(combined_df.columns).index(col_id_name) + 1
            
        for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if c_idx == order_item_col_idx and r_idx > 1:
                    cell.value = str(val).strip()
                    cell.data_type = 's'
                    cell.number_format = '@'
                else:
                    cell.value = val
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        wb.save(temp_path)
        
    except Exception as e:
        return jsonify({'error': f'Error generating Excel file: {str(e)}'}), 500
        
    return jsonify({
        'message': 'Files successfully merged and processed!',
        'total_orders': len(combined_df),
        'columns': columns,
        'preview': preview_data
    })

# API endpoint to check if mapping file is saved
@app.route('/api/mapping-status', methods=['GET'])
def get_mapping_status():
    mapping_df = load_saved_mapping()
    if mapping_df is not None:
        return jsonify({
            'loaded': True,
            'rules_count': len(mapping_df)
        })
    return jsonify({'loaded': False})

@app.route('/api/rename', methods=['POST'])
def rename_files():
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
        
    uploaded_files = request.files.getlist('files[]')
    option = request.form.get('option', 'no')
    
    print(f"[DEBUG rename] Received option: {option}", flush=True)
    
    if not uploaded_files or len(uploaded_files) == 0 or uploaded_files[0].filename == '':
        return jsonify({'error': 'No files uploaded.'}), 400
        
    # Read files into memory dictionaries to separate mapping from rename sheets
    mapping_file_data = None
    rename_files_list = []
    
    for f in uploaded_files:
        if not (f.filename.endswith('.xlsx') or f.filename.endswith('.xls') or f.filename.endswith('.csv')):
            continue
        try:
            f_bytes = f.read()
            if is_mapping_file(f.filename, f_bytes):
                mapping_file_data = {
                    'filename': f.filename,
                    'bytes': f_bytes
                }
            else:
                rename_files_list.append({
                    'filename': f.filename,
                    'bytes': f_bytes
                })
        except Exception as e:
            return jsonify({'error': f'Failed loading {f.filename}: {str(e)}'}), 400
            
    # Process Brand Mapping
    mapping_df = None
    mapping_name_detected = None
    
    if mapping_file_data:
        mapping_df = parse_mapping_df(mapping_file_data['bytes'], mapping_file_data['filename'])
        if mapping_df is not None:
            mapping_name_detected = mapping_file_data['filename']
            # Save it to temp directory to persist for future sessions
            ext = os.path.splitext(mapping_file_data['filename'])[1]
            os.makedirs('temp', exist_ok=True)
            for old_ext in ['.xlsx', '.xls', '.csv']:
                old_path = os.path.join('temp', f'arrange_mapping{old_ext}')
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except Exception:
                        pass
            dest_path = os.path.join('temp', f'arrange_mapping{ext}')
            with open(dest_path, 'wb') as df_out:
                df_out.write(mapping_file_data['bytes'])
    else:
        # Load previously saved mapping file
        mapping_df = load_saved_mapping()
        
    if option == 'yes' and mapping_df is None:
        return jsonify({
            'error': 'Option A requires the ARRANGE mapping sheet. Please upload your mapping file along with files to rename.'
        }), 400

    if not rename_files_list:
        if mapping_file_data:
            return jsonify({
                'message': 'ARRANGE mapping sheet uploaded and saved successfully!',
                'type': 'mapping_only',
                'rules_count': len(mapping_df) if mapping_df is not None else 0,
                'filename': mapping_file_data['filename']
            })
        else:
            return jsonify({'error': 'No files to rename were uploaded.'}), 400

    processed_files = []
    renamed_log = []
    
    for rf in rename_files_list:
        fname = rf['filename']
        fbytes = rf['bytes']
        
        rename_code = ""
        if option == 'yes':
            rename_code = find_rename_code_option_a_from_bytes(fbytes, fname, mapping_df)
        else:
            rename_code = find_rename_code_option_b_from_bytes(fbytes, fname)
            
        rename_code = str(rename_code).strip()
        print(f"[DEBUG rename] File: {fname}, Extracted rename code: {rename_code}", flush=True)
        
        # Get extension
        ext_idx = fname.rfind('.')
        if ext_idx != -1:
            name_part = fname[:ext_idx]
            ext_part = fname[ext_idx:]
        else:
            name_part = fname
            ext_part = ""
            
        if rename_code:
            new_fname = f"{rename_code}-{name_part}{ext_part}"
        else:
            new_fname = fname
            
        processed_files.append((new_fname, fbytes))
        renamed_log.append({
            'original': fname,
            'renamed': new_fname,
            'code': rename_code or 'None'
        })
        
    # Handle duplicates in names within the batch
    final_files = []
    seen_names = {}
    
    for filename, data in processed_files:
        if filename in seen_names:
            seen_names[filename] += 1
            ext_idx = filename.rfind('.')
            if ext_idx != -1:
                name_part = filename[:ext_idx]
                ext_part = filename[ext_idx:]
            else:
                name_part = filename
                ext_part = ""
            new_filename = f"{name_part} ({seen_names[filename]}){ext_part}"
        else:
            seen_names[filename] = 0
            new_filename = filename
        final_files.append((new_filename, data))

    # Package output
    if len(final_files) == 1:
        # Single file
        single_name, single_data = final_files[0]
        os.makedirs('temp', exist_ok=True)
        temp_rename_path = os.path.join('temp', 'renamed_output_file')
        with open(temp_rename_path, 'wb') as f:
            f.write(single_data)
            
        return jsonify({
            'message': 'File successfully renamed!',
            'type': 'single',
            'filename': single_name,
            'log': renamed_log,
            'mapping_detected': mapping_name_detected
        })
    else:
        # ZIP
        os.makedirs('temp', exist_ok=True)
        zip_path = os.path.join('temp', 'renamed_output_zip.zip')
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, data in final_files:
                zip_file.writestr(filename, data)
                
        return jsonify({
            'message': f'Successfully renamed {len(final_files)} files!',
            'type': 'zip',
            'filename': 'Renamed_Files.zip',
            'log': renamed_log,
            'mapping_detected': mapping_name_detected
        })

@app.route('/api/download-renamed', methods=['GET'])
def download_renamed():
    download_type = request.args.get('type', 'single')
    custom_filename = request.args.get('filename', 'renamed_output')
    
    if download_type == 'zip':
        temp_path = os.path.join('temp', 'renamed_output_zip.zip')
        mimetype = 'application/zip'
    else:
        temp_path = os.path.join('temp', 'renamed_output_file')
        if custom_filename.endswith('.csv'):
            mimetype = 'text/csv'
        else:
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    if not os.path.exists(temp_path):
        return jsonify({'error': 'Renamed file not found. Please upload and rename again.'}), 400
        
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=custom_filename,
        mimetype=mimetype
    )

@app.route('/api/download', methods=['GET'])
def download_merged_file():
    temp_path = os.path.join('temp', 'merged_output.xlsx')
    if not os.path.exists(temp_path):
        return jsonify({'error': 'Merged file not found. Please upload files again.'}), 400
    return send_file(
        temp_path,
        as_attachment=True,
        download_name='Flipkart_Merged_Orders.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def run_split_logic_from_bytes(file_bytes, filename, option, col_choice=None):
    # Load sheet headerless to preserve structure of cover and header rows
    df = load_any_sheet_to_dataframe(file_bytes, filename, header=None)
    
    # Determine parameters based on option
    if option == "1":
        name_suffix = "-FLIPKART"
        filter_field_idx = 3  # Column D (4)
        data_start_row_idx = 2  # 0-based index for Row 3
        header_rows_count = 2
    elif option == "2":
        name_suffix = " DETAILS SHEET FLIPKART"
        filter_field_idx = 3  # Column D (4)
        data_start_row_idx = 2
        header_rows_count = 2
    elif option == "3":
        name_suffix = " SUMMARY SHEET FLIPKART"
        filter_field_idx = 6  # Column G (7)
        data_start_row_idx = 2
        header_rows_count = 2
    elif option == "4":
        name_suffix = ""
        filter_field_idx = 0  # Column A (1)
        data_start_row_idx = 1  # 0-based index for Row 2
        header_rows_count = 1
    else:
        raise ValueError("Invalid option selected")
        
    if df.shape[0] <= data_start_row_idx:
        print(f"[DEBUG Split] DataFrame too small: rows = {df.shape[0]}", flush=True)
        return []
        
    # Headers are rows before data_start_row_idx
    header_df = df.iloc[:data_start_row_idx]
    data_df = df.iloc[data_start_row_idx:]
    
    if filter_field_idx >= df.shape[1]:
        # Fallback if indices mismatch
        filter_field_idx = 0
        print(f"[DEBUG Split] Column choice {filter_field_idx} out of range, falling back to index 0", flush=True)
        
    # Get unique filter values
    filter_values = data_df.iloc[:, filter_field_idx].dropna().astype(str).str.strip()
    unique_keys = [k for k in filter_values.unique() if k and k.lower() != 'nan' and k.lower() != 'none']
    
    split_files = []
    import datetime
    timestamp_str = datetime.datetime.now().strftime("%d-%m-%Y %H-%M-%S")
    file_counter = 1
    
    for key in unique_keys:
        # Filter matching rows
        matched_rows = data_df[data_df.iloc[:, filter_field_idx].astype(str).str.strip() == key]
        if matched_rows.empty:
            continue
            
        # Combine header and matched rows
        combined_df = pd.concat([header_df, matched_rows], ignore_index=True)
        
        # Determine final name
        if option == "4":
            first_num = key.split('-')[0].strip()
            final_name = f"{first_num}-Tax-{key}-MYNTYRA"
        else:
            final_name = f"{key}{name_suffix}"
            
        # Clean invalid filename characters
        for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
            final_name = final_name.replace(char, '')
        final_name = final_name.strip()
        
        # Append date stamp and counter format: whName & " " & dtStamp & "_" & Format(fileCounter, "00") & ".xlsx"
        final_filename = f"{final_name} {timestamp_str}_{file_counter:02d}.xlsx"
        
        # Build Excel sheet
        out_io = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=False), start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if val == 'None' or pd.isna(val):
                    cell.value = ""
                else:
                    cell.value = val
                    
        # Apply standard auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        wb.save(out_io)
        out_io.seek(0)
        
        split_files.append((final_filename, out_io.getvalue()))
        file_counter += 1
        
    return split_files

@app.route('/api/split', methods=['POST'])
def split_file_route():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
        
    file = request.files['file']
    option = request.form.get('option', '1')
    col_choice = request.form.get('col_choice', 'G')
    
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
        
    try:
        file_bytes = file.read()
        
        # Call split logic
        split_files = run_split_logic_from_bytes(file_bytes, file.filename, option, col_choice)
        
        if not split_files:
            return jsonify({'error': 'No data found to split or sheet is empty'}), 400
            
        # Save split files to a ZIP archive in temp/
        os.makedirs('temp', exist_ok=True)
        zip_path = os.path.join('temp', 'split_output.zip')
        
        # Remove old zip if exists
        if os.path.exists(zip_path):
            try:
                os.remove(zip_path)
            except Exception:
                pass
                
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for s_name, s_bytes in split_files:
                zip_file.writestr(s_name, s_bytes)
                
        # Generate log output
        log_entries = []
        for idx, (s_name, _) in enumerate(split_files, start=1):
            key_extracted = s_name
            if option == "4":
                if "-Tax-" in s_name:
                    parts = s_name.split("-Tax-")
                    if len(parts) >= 2:
                        key_part = parts[1]
                        if "-MYNTYRA" in key_part:
                            key_extracted = key_part.split("-MYNTYRA")[0].strip()
            else:
                suffix = "-FLIPKART" if option == "1" else (" DETAILS SHEET FLIPKART" if option == "2" else " SUMMARY SHEET FLIPKART")
                if suffix in s_name:
                    key_extracted = s_name.split(suffix)[0].strip()
                    
            log_entries.append({
                'index': f"{idx:02d}",
                'filename': s_name,
                'key': key_extracted
            })
            
        import datetime
        safe_orig_name = "".join(c for c in os.path.splitext(file.filename)[0] if c.isalnum() or c in ['_', '-']).strip()
        zip_filename = f"Split_{safe_orig_name}_{datetime.datetime.now().strftime('%d-%m-%Y')}.zip"
        
        return jsonify({
            'message': 'Spreadsheet split successfully!',
            'files_count': len(split_files),
            'zip_filename': zip_filename,
            'log': log_entries
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed splitting file: {str(e)}'}), 500

@app.route('/api/create-folder', methods=['POST'])
def create_folder():
    try:
        if 'files[]' not in request.files:
            return jsonify({'error': 'No files uploaded'}), 400
            
        uploaded_files = request.files.getlist('files[]')
        if not uploaded_files or len(uploaded_files) < 2:
            return jsonify({'error': 'Please upload at least 2 files (merged file + other files)'}), 400
            
        from werkzeug.utils import secure_filename
        
        # Classify files
        merged_file = None
        other_files = []
        
        for file in uploaded_files:
            if not file.filename:
                continue
            if 'FLIPKART_MERGED_ORDERS' in file.filename.upper():
                merged_file = file
            else:
                other_files.append(file)
                
        if not merged_file:
            return jsonify({'error': 'No file containing "FLIPKART_MERGED_ORDERS" in its name was found.'}), 400
            
        if not other_files:
            return jsonify({'error': 'No other prefix files were found.'}), 400
            
        # We will work in a temporary folder
        import shutil
        import tempfile
        
        # Create temp work directory
        temp_work_dir = tempfile.mkdtemp(dir='temp')
        
        # Save merged file bytes to reuse
        merged_bytes = merged_file.read()
        merged_filename = secure_filename(merged_file.filename)
        
        # Map to track operations
        # prefix -> list of other filenames
        prefix_groups = {}
        
        for file in other_files:
            filename = file.filename
            # Extract prefix before the first '-'
            if '-' in filename:
                prefix = filename.split('-', 1)[0].strip()
                if prefix:
                    if prefix not in prefix_groups:
                        prefix_groups[prefix] = []
                    
                    # Create folder for prefix if it doesn't exist
                    dest_folder_path = os.path.join(temp_work_dir, prefix)
                    os.makedirs(dest_folder_path, exist_ok=True)
                    
                    # Copy merged file
                    merged_dest_path = os.path.join(dest_folder_path, merged_filename)
                    if not os.path.exists(merged_dest_path):
                        with open(merged_dest_path, 'wb') as f_out:
                            f_out.write(merged_bytes)
                            
                    # Save the other file directly
                    other_dest_path = os.path.join(dest_folder_path, secure_filename(filename))
                    # Reset pointer just in case and save bytes
                    file.seek(0)
                    with open(other_dest_path, 'wb') as f_other:
                        f_other.write(file.read())
                    
                    prefix_groups[prefix].append(filename)
                    
        if not prefix_groups:
            # Clean up and return error
            shutil.rmtree(temp_work_dir, ignore_errors=True)
            return jsonify({'error': 'No files with valid prefix (e.g. "101-") found to group.'}), 400
            
        # Sort prefixes numerically if numeric, else alphabetically
        numeric_prefixes = []
        alpha_prefixes = []
        for p in prefix_groups.keys():
            if p.isdigit():
                numeric_prefixes.append(int(p))
            else:
                alpha_prefixes.append(p)
                
        numeric_prefixes.sort()
        alpha_prefixes.sort(key=str.upper)
        
        sorted_prefixes = [str(n) for n in numeric_prefixes] + alpha_prefixes
        
        # ZIP filename pattern: firstPrefix-lastPrefix.zip
        first_p = sorted_prefixes[0]
        last_p = sorted_prefixes[-1]
        zip_filename = f"{first_p}-{last_p}.zip"
        
        # Create output ZIP archive
        temp_zip_path = os.path.join('temp', 'folder_output.zip')
        if os.path.exists(temp_zip_path):
            os.remove(temp_zip_path)
            
        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for p in sorted_prefixes:
                p_dir = os.path.join(temp_work_dir, p)
                for root, dirs, files in os.walk(p_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        # Archive path: prefix/file
                        archive_path = os.path.relpath(file_path, temp_work_dir)
                        z.write(file_path, archive_path)
                        
        # Clean up temp folder
        shutil.rmtree(temp_work_dir, ignore_errors=True)
        
        # Prepare response logs
        log_entries = []
        for p in sorted_prefixes:
            log_entries.append({
                'folder': p,
                'copied_merged': merged_filename,
                'moved_files': prefix_groups[p]
            })
            
        return jsonify({
            'message': 'Folders created and zipped successfully!',
            'folders_count': len(sorted_prefixes),
            'zip_filename': zip_filename,
            'log': log_entries
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed grouping folders: {str(e)}'}), 500

@app.route('/api/download-folder-zip', methods=['GET'])
def download_folder_zip():
    custom_filename = request.args.get('filename', 'Grouped_Folders.zip')
    temp_path = os.path.join('temp', 'folder_output.zip')
    
    if not os.path.exists(temp_path):
        return jsonify({'error': 'Grouped ZIP archive not found. Please upload and process again.'}), 400
        
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=custom_filename,
        mimetype='application/zip'
    )

def load_file_to_openpyxl_workbook(file_bytes, filename):
    filetype = detect_file_type(file_bytes, filename)
    if is_html_bytes(file_bytes):
        filetype = 'html'
        
    if filetype == 'xlsx':
        try:
            return load_workbook(io.BytesIO(file_bytes))
        except Exception:
            pass
            
    try:
        df = load_any_sheet_to_dataframe(file_bytes, filename, header=None)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        return wb
    except Exception as e:
        raise ValueError(f"Could not load file {filename} to openpyxl workbook: {str(e)}")

def process_single_party(od_bytes, od_filename, dt_bytes, dt_filename, details_bytes, details_filename, dest_folder):
    # Load Details file to extract Invoice Numbers from Column B starting at row 3
    df_details = load_any_sheet_to_dataframe(details_bytes, details_filename, header=None)
    
    details_invoices = []
    if df_details.shape[0] > 2:
        details_invoices = df_details.iloc[2:, 1].dropna().astype(str).str.strip().tolist()
        details_invoices = [inv.replace('`', '').replace('"', '').strip() for inv in details_invoices if inv]
        
    # Load DT file
    wb_dt = load_file_to_openpyxl_workbook(dt_bytes, dt_filename)
    ws_dt = wb_dt.active
    
    # Step A: Delete matching rows
    deleted_count = 0
    for r in range(ws_dt.max_row, 1, -1):
        val = ws_dt.cell(row=r, column=7).value
        if val:
            val_clean = str(val).replace('`', '').replace('"', '').strip()
            if val_clean in details_invoices:
                ws_dt.delete_rows(r, 1)
                deleted_count += 1
                
    # Step B: Prefix and suffix calculation
    file_prefix = "101"
    if '-' in dt_filename:
        file_prefix = dt_filename.split('-', 1)[0].strip()
        
    suffixes = []
    invoice_prefix = None
    for r in range(2, ws_dt.max_row + 1):
        val = ws_dt.cell(row=r, column=7).value
        if val:
            val_str = str(val).strip()
            if '-' in val_str:
                parts = val_str.rsplit('-', 1)
                p_part = parts[0].replace('`', '').replace('"', '').strip()
                s_part = parts[1].strip()
                if not invoice_prefix:
                    invoice_prefix = p_part
                if s_part.isdigit():
                    suffixes.append(int(s_part))
            else:
                val_clean = val_str.replace('`', '').replace('"', '').strip()
                if val_clean.isdigit():
                    suffixes.append(int(val_clean))
                if not invoice_prefix:
                    invoice_prefix = val_clean
                    
    if suffixes:
        min_s = min(suffixes)
        max_s = max(suffixes)
        suffix_range = f"{min_s}-{max_s}" if min_s != max_s else f"{min_s}"
    else:
        suffix_range = "empty"
        
    if not invoice_prefix:
        invoice_prefix = "FK"
        
    dt_new_name = f"{file_prefix}-({suffix_range})-DT.xlsx"
    od_new_name = f"{file_prefix}-({invoice_prefix}-{suffix_range})-OD.xlsx"
    pr_new_name = f"{file_prefix}-({invoice_prefix}-{suffix_range})-PR.xlsx"
    
    # Step C: Collect clean suborder numbers
    suborders_in_dt = set()
    for r in range(2, ws_dt.max_row + 1):
        val = ws_dt.cell(row=r, column=5).value
        if val:
            val_clean = str(val).replace('`', '').replace('"', '').strip()
            if val_clean:
                suborders_in_dt.add(val_clean)
                
    # Load OD file
    wb_od = load_file_to_openpyxl_workbook(od_bytes, od_filename)
    ws_od = wb_od.active
    
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    
    yellow_fill = PatternFill(start_color="FFFFB4", end_color="FFFFB4", fill_type="solid")
    
    wb_matched_od = Workbook()
    ws_matched_od = wb_matched_od.active
    ws_matched_od.title = "Matched Orders"
    
    headers_od = [cell.value for cell in ws_od[1]]
    ws_matched_od.append(headers_od)
    
    matched_od_count = 0
    for r in range(2, ws_od.max_row + 1):
        val = ws_od.cell(row=r, column=1).value
        if val:
            val_clean = str(val).replace('`', '').replace('"', '').strip()
            if val_clean in suborders_in_dt:
                matched_od_count += 1
                row_vals = [cell.value for cell in ws_od[r]]
                ws_matched_od.append(row_vals)
                for col in range(1, ws_od.max_column + 1):
                    ws_od.cell(row=r, column=col).fill = yellow_fill
                    
    # Step D: GST Not Applicable Export
    wb_gst = Workbook()
    ws_gst = wb_gst.active
    ws_gst.title = "GST NOT APPLICABLE"
    
    gst_headers = ["EE Invoice No", "Order Status", "Invoice Date", "Item Quantity", "Selling Price", "Item Price(Excluding Tax)"]
    ws_gst.append(gst_headers)
    
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_idx in range(1, 7):
        cell = ws_gst.cell(row=1, column=col_idx)
        cell.fill = black_fill
        cell.font = white_bold_font
        
    light_green_fill = PatternFill(start_color="C8FFC8", end_color="C8FFC8", fill_type="solid")
    ap_green_fill = PatternFill(start_color="B4F0B4", end_color="B4F0B4", fill_type="solid")
    
    shade_index = 1
    new_row_gst = 2
    gst_exported_count = 0
    
    for r in range(2, ws_dt.max_row + 1):
        invoice_val = ws_dt.cell(row=r, column=7).value
        ap_val = ws_dt.cell(row=r, column=42).value
        
        if invoice_val and (not ap_val or str(ap_val).strip() == ""):
            val_g = ws_dt.cell(row=r, column=7).value
            val_i = ws_dt.cell(row=r, column=9).value
            val_m = ws_dt.cell(row=r, column=13).value
            val_r = ws_dt.cell(row=r, column=18).value
            val_av = ws_dt.cell(row=r, column=48).value
            val_ax = ws_dt.cell(row=r, column=50).value
            
            ws_gst.append([val_g, val_i, val_m, val_r, val_av, val_ax])
            
            Rc = 170 + ((shade_index * 37) % 80)
            Gc = 170 + ((shade_index * 67) % 80)
            Bc = 170 + ((shade_index * 97) % 80)
            color_hex = f"{Rc:02X}{Gc:02X}{Bc:02X}"
            row_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            for c_idx in range(1, 7):
                ws_gst.cell(row=new_row_gst, column=c_idx).fill = row_fill
                
            for c_col in [7, 9, 13, 18, 48, 50]:
                ws_dt.cell(row=r, column=c_col).fill = light_green_fill
            ws_dt.cell(row=r, column=42).fill = ap_green_fill
            
            shade_index += 1
            new_row_gst += 1
            gst_exported_count += 1
            
    # Duplicates check
    invoice_counts = {}
    for r in range(2, ws_dt.max_row + 1):
        val = ws_dt.cell(row=r, column=7).value
        if val:
            val_clean = str(val).strip()
            invoice_counts[val_clean] = invoice_counts.get(val_clean, 0) + 1
            
    duplicates = {k: v for k, v in invoice_counts.items() if v > 1}
    
    # State mapping dictionary
    STATE_TO_CODE = {
        'ANDHRA PRADESH': 'AP', 'ARUNACHAL PRADESH': 'AR', 'ASSAM': 'AS', 'BIHAR': 'BR',
        'CHHATTISGARH': 'CG', 'GOA': 'GA', 'GUJARAT': 'GJ', 'HARYANA': 'HR',
        'HIMACHAL PRADESH': 'HP', 'JHARKHAND': 'JH', 'KARNATAKA': 'KA', 'KERALA': 'KL',
        'MADHYA PRADESH': 'MP', 'MAHARASHTRA': 'MH', 'MANIPUR': 'MN', 'MEGHALAYA': 'ML',
        'MIZORAM': 'MZ', 'NAGALAND': 'NL', 'ODISHA': 'OD', 'ORISSA': 'OD', 'PUNJAB': 'PB',
        'RAJASTHAN': 'RJ', 'SIKKIM': 'SK', 'TAMIL NADU': 'TN', 'TELANGANA': 'TG',
        'TRIPURA': 'TR', 'UTTAR PRADESH': 'UP', 'UTTARAKHAND': 'UK', 'UTTARANCHAL': 'UK',
        'WEST BENGAL': 'WB', 'ANDAMAN AND NICOBAR ISLANDS': 'AN', 'CHANDIGARH': 'CH',
        'DADRA AND NAGAR HAVELI': 'DN', 'DAMAN AND DIU': 'DD', 'DELHI': 'DL',
        'JAMMU AND KASHMIR': 'JK', 'JAMMU & KASHMIR': 'JK', 'LAKSHADWEEP': 'LD',
        'PUDUCHERRY': 'PY', 'PONDICHERRY': 'PY', 'LADAKH': 'LA'
    }

    # Step E: Create PR Workbook
    wb_pr = Workbook()
    ws_pr = wb_pr.active
    ws_pr.title = "PR_Report"

    # Headers
    pr_headers = [
        "Order ID", "Invoice ID", "New Invoice ID", "Invoice Reference Number (IRN)",
        "Shipment date", "Invoice date", "GST ID", "FSN CODE", "SKU", "Item Title",
        "Quantity", "Item Cost", "GST Rate", "CESS Rate", "HSN", "Warehouse Code/Name",
        "Status", "state code", "Promotion Discount(Excluding Tax)"
    ]
    ws_pr.append(pr_headers)

    # Build lookup from OD file
    od_lookup = {}
    for r_od in range(2, ws_od.max_row + 1):
        item_id = ws_od.cell(row=r_od, column=1).value
        if item_id:
            clean_item_id = str(item_id).replace('`', '').replace('"', '').strip()
            fsn_val = ws_od.cell(row=r_od, column=9).value
            sku_val = ws_od.cell(row=r_od, column=8).value
            title_val = ws_od.cell(row=r_od, column=10).value
            od_lookup[clean_item_id] = (fsn_val, sku_val, title_val)

    # Map remaining DT rows to PR workbook
    pr_row = 2
    for r_dt in range(2, ws_dt.max_row + 1):
        order_id_raw = ws_dt.cell(row=r_dt, column=5).value
        invoice_id_raw = ws_dt.cell(row=r_dt, column=4).value
        new_invoice_id_raw = ws_dt.cell(row=r_dt, column=7).value
        shipment_date_raw = ws_dt.cell(row=r_dt, column=12).value
        invoice_date_raw = ws_dt.cell(row=r_dt, column=13).value
        gst_id_raw = ws_dt.cell(row=r_dt, column=2).value
        quantity_raw = ws_dt.cell(row=r_dt, column=18).value
        item_cost_raw = ws_dt.cell(row=r_dt, column=50).value
        gst_rate_raw = ws_dt.cell(row=r_dt, column=42).value
        hsn_raw = ws_dt.cell(row=r_dt, column=26).value
        wh_raw = ws_dt.cell(row=r_dt, column=1).value
        state_raw = ws_dt.cell(row=r_dt, column=40).value
        promotion_discount_raw = ws_dt.cell(row=r_dt, column=54).value
        
        # Format GST Rate with %
        gst_rate_formatted = ""
        if gst_rate_raw is not None:
            gst_rate_formatted = str(gst_rate_raw).strip()
            if gst_rate_formatted and not gst_rate_formatted.endswith('%'):
                gst_rate_formatted += "%"
                
        # Map State code
        state_code_formatted = ""
        if state_raw:
            state_str = str(state_raw).strip().upper()
            state_clean = state_str.replace('&', 'AND')
            state_clean = " ".join(state_clean.split())
            if state_str in STATE_TO_CODE:
                state_code_formatted = STATE_TO_CODE[state_str]
            elif state_clean in STATE_TO_CODE:
                state_code_formatted = STATE_TO_CODE[state_clean]
            else:
                state_code_formatted = state_raw
            
        # Write mapped cells to PR
        ws_pr.cell(row=pr_row, column=1, value=order_id_raw)
        ws_pr.cell(row=pr_row, column=2, value=invoice_id_raw)
        ws_pr.cell(row=pr_row, column=3, value=new_invoice_id_raw)
        ws_pr.cell(row=pr_row, column=5, value=shipment_date_raw)
        ws_pr.cell(row=pr_row, column=6, value=invoice_date_raw)
        ws_pr.cell(row=pr_row, column=7, value=gst_id_raw)
        ws_pr.cell(row=pr_row, column=11, value=quantity_raw)
        ws_pr.cell(row=pr_row, column=12, value=item_cost_raw)
        ws_pr.cell(row=pr_row, column=13, value=gst_rate_formatted)
        ws_pr.cell(row=pr_row, column=15, value=hsn_raw)
        ws_pr.cell(row=pr_row, column=16, value=wh_raw)
        ws_pr.cell(row=pr_row, column=18, value=state_code_formatted)
        ws_pr.cell(row=pr_row, column=19, value=promotion_discount_raw)
        
        # Lookup OD details
        if order_id_raw:
            clean_order_id = str(order_id_raw).replace('`', '').replace('"', '').strip()
            if clean_order_id in od_lookup:
                sku_id_val, sku_val, title_val = od_lookup[clean_order_id]
                ws_pr.cell(row=pr_row, column=8, value=sku_id_val)
                ws_pr.cell(row=pr_row, column=9, value=sku_val)
                ws_pr.cell(row=pr_row, column=10, value=title_val)
                
        pr_row += 1

    # Create nested subfolder name (e.g. 101-(FK27S101-1-9))
    subfolder_name = f"{file_prefix}-({invoice_prefix}-{suffix_range})"
    actual_dest_folder = os.path.join(dest_folder, subfolder_name)
    os.makedirs(actual_dest_folder, exist_ok=True)

    # Save DT file
    dt_out_path = os.path.join(actual_dest_folder, dt_new_name)
    for col in ws_dt.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_dt.column_dimensions[col_letter].width = max(max_len + 3, 10)
    wb_dt.save(dt_out_path)
    
    # Save OD file
    od_out_path = os.path.join(actual_dest_folder, od_new_name)
    for col in ws_matched_od.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_matched_od.column_dimensions[col_letter].width = max(max_len + 3, 10)
    wb_matched_od.save(od_out_path)

    # Save PR file
    pr_out_path = os.path.join(actual_dest_folder, pr_new_name)
    for col in ws_pr.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_pr.column_dimensions[col_letter].width = max(max_len + 3, 10)
    wb_pr.save(pr_out_path)
    
    gst_created = False
    if gst_exported_count > 0:
        for col in ws_gst.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_gst.column_dimensions[col_letter].width = max(max_len + 3, 10)
        wb_gst.save(os.path.join(actual_dest_folder, "GST NOT APPLICABLE.xlsx"))
        gst_created = True
        
    dup_created = False
    if duplicates:
        wb_dup = Workbook()
        ws_dup = wb_dup.active
        ws_dup.title = "Duplicates"
        ws_dup.cell(row=1, column=1, value="DUPLICATE INVOICE").font = Font(bold=True)
        ws_dup.cell(row=1, column=2, value="COUNT").font = Font(bold=True)
        
        d_row = 2
        for k, v in duplicates.items():
            ws_dup.cell(row=d_row, column=1, value=k)
            ws_dup.cell(row=d_row, column=2, value=v)
            d_row += 1
            
        ws_dup.column_dimensions['A'].width = 25
        ws_dup.column_dimensions['B'].width = 15
        wb_dup.save(os.path.join(actual_dest_folder, "2 MORE INVOICE.xlsx"))
        dup_created = True
        
    return {
        'prefix': file_prefix,
        'invoice_prefix': invoice_prefix,
        'suffix_range': suffix_range,
        'subfolder_name': subfolder_name,
        'deleted_count': deleted_count,
        'matched_od_count': matched_od_count,
        'gst_created': gst_created,
        'gst_exported_count': gst_exported_count,
        'dup_created': dup_created,
        'duplicates_count': len(duplicates),
        'dt_new_name': dt_new_name,
        'od_new_name': od_new_name,
        'pr_new_name': pr_new_name
    }

@app.route('/api/invoice-arrange', methods=['POST'])
def invoice_arrange():
    try:
        import tempfile
        import shutil
        import zipfile
        
        # We check if a zipfile is uploaded
        is_zip_upload = 'zipfile' in request.files
        
        temp_work_dir = tempfile.mkdtemp(dir='temp')
        log_summary = []
        processed_folders = []
        summary_records = []
        
        if is_zip_upload:
            zip_file = request.files['zipfile']
            zip_bytes = zip_file.read()
            
            # Temporary extraction directory
            extract_dir = tempfile.mkdtemp(dir='temp')
            with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
                z.extractall(extract_dir)
                
            # Scan extract_dir for folders containing OD, DT, and Details files
            folders_to_process = {}
            for root, dirs, files in os.walk(extract_dir):
                valid_files = [f for f in files if not f.startswith('.') and not f.startswith('__')]
                if not valid_files:
                    continue
                rel_path = os.path.relpath(root, extract_dir)
                if rel_path == '.':
                    folder_name = "arranged_files"
                else:
                    folder_name = os.path.basename(rel_path)
                    
                folders_to_process[root] = {
                    'files': valid_files,
                    'folder_name': folder_name
                }
                
            # Process each found folder
            total_files_count = 0
            for folder_path, info in folders_to_process.items():
                files = info['files']
                folder_name = info['folder_name']
                
                od_path = None
                dt_path = None
                details_path = None
                
                for f in files:
                    f_upper = f.upper()
                    full_p = os.path.join(folder_path, f)
                    if 'FLIPKART_MERGED_ORDERS' in f_upper:
                        od_path = full_p
                    elif 'TAXREPORTDATA' in f_upper:
                        dt_path = full_p
                    else:
                        details_path = full_p
                        
                if not od_path or not dt_path or not details_path:
                    continue
                    
                # Read files
                with open(od_path, 'rb') as f_od:
                    od_bytes = f_od.read()
                with open(dt_path, 'rb') as f_dt:
                    dt_bytes = f_dt.read()
                with open(details_path, 'rb') as f_det:
                    details_bytes = f_det.read()
                    
                # Destination folder: temp_work_dir/folder_name
                dest_dir = os.path.join(temp_work_dir, folder_name)
                
                # Process single folder
                res = process_single_party(
                    od_bytes, os.path.basename(od_path),
                    dt_bytes, os.path.basename(dt_path),
                    details_bytes, os.path.basename(details_path),
                    dest_dir
                )
                
                processed_folders.append(res['prefix'])
                summary_records.append({
                    'prefix': res['prefix'],
                    'invoice_range_str': f"{res['invoice_prefix']}-{res['suffix_range']}"
                })
                
                # Logs for this folder
                log_summary.append({
                    'operation': f"[{res['prefix']}] Deleted Matches",
                    'value': f"{res['deleted_count']} rows",
                    'status': 'Success'
                })
                log_summary.append({
                    'operation': f"[{res['prefix']}] PR & OD Matching",
                    'value': f"{res['matched_od_count']} highlighted",
                    'status': 'Success'
                })
                
                folder_files_count = 3 + (1 if res['gst_created'] else 0) + (1 if res['dup_created'] else 0)
                total_files_count += folder_files_count
                
            shutil.rmtree(extract_dir, ignore_errors=True)
            
            if not processed_folders:
                shutil.rmtree(temp_work_dir, ignore_errors=True)
                return jsonify({'error': 'No valid folders containing OD, DT, and Details files were found in the zip.'}), 400
                
            # Sorted prefixes to generate the zip file name
            valid_prefixes = [p for p in processed_folders if p]
            try:
                sorted_prefixes = sorted(valid_prefixes, key=lambda x: int(x) if x.isdigit() else x)
            except Exception:
                sorted_prefixes = sorted(valid_prefixes)
                
            first_prefix = sorted_prefixes[0]
            last_prefix = sorted_prefixes[-1]
            zip_display_name = f"{first_prefix}-{last_prefix}-arranged.zip" if first_prefix != last_prefix else f"{first_prefix}-arranged.zip"
            
        else:
            # Single-party 3-file upload
            if 'OD' not in request.files or 'DT' not in request.files or 'Details' not in request.files:
                shutil.rmtree(temp_work_dir, ignore_errors=True)
                return jsonify({'error': 'Missing files. Please upload exactly 3 files: OD, DT, and Details.'}), 400
                
            od_file = request.files['OD']
            dt_file = request.files['DT']
            details_file = request.files['Details']
            
            od_bytes = od_file.read()
            dt_bytes = dt_file.read()
            details_bytes = details_file.read()
            
            # Extract prefix from DT name or set default
            prefix_folder = "arranged"
            if '-' in dt_file.filename:
                prefix_folder = dt_file.filename.split('-', 1)[0].strip()
                
            dest_dir = os.path.join(temp_work_dir, prefix_folder)
            
            res = process_single_party(
                od_bytes, od_file.filename,
                dt_bytes, dt_file.filename,
                details_bytes, details_file.filename,
                dest_dir
            )
            
            summary_records.append({
                'prefix': res['prefix'],
                'invoice_range_str': f"{res['invoice_prefix']}-{res['suffix_range']}"
            })
            
            log_summary = [
                {'operation': 'Matched Rows Deleted', 'value': str(res['deleted_count']), 'status': 'Success (DT rows cleared)'},
                {'operation': 'Output DT File Created', 'value': f"{prefix_folder}/{res['subfolder_name']}/{res['dt_new_name']}", 'status': 'Success'},
                {'operation': 'Matched Suborders Highlighted', 'value': str(res['matched_od_count']), 'status': 'Success'},
                {'operation': 'Output OD File Created', 'value': f"{prefix_folder}/{res['subfolder_name']}/{res['od_new_name']}", 'status': 'Success'},
                {'operation': 'Output PR File Created', 'value': f"{prefix_folder}/{res['subfolder_name']}/{res['pr_new_name']}", 'status': 'Success'},
                {'operation': 'GST Not Applicable Export', 'value': f"{prefix_folder}/{res['subfolder_name']}/GST NOT APPLICABLE.xlsx" if res['gst_created'] else 'None', 'status': f"Exported {res['gst_exported_count']} rows" if res['gst_created'] else 'No matching rows'},
                {'operation': 'Duplicate Invoice Check', 'value': f"{prefix_folder}/{res['subfolder_name']}/2 MORE INVOICE.xlsx" if res['dup_created'] else 'None', 'status': f"Found {res['duplicates_count']} duplicates" if res['dup_created'] else 'No duplicates found'}
            ]
            
            total_files_count = 3 + (1 if res['gst_created'] else 0) + (1 if res['dup_created'] else 0)
            zip_display_name = f"{res['prefix']}-({res['invoice_prefix']}-{res['suffix_range']})-arranged.zip"
            
        # Fetch party name mapping from Apps Script
        party_mapping = {}
        try:
            import requests
            r = requests.get(APPS_SCRIPT_URL, allow_redirects=True, timeout=10)
            if r.status_code == 200:
                for item in r.json():
                    code = str(item.get('CODE', '')).strip()
                    party_code = str(item.get('PARTY CODE', '')).strip()
                    if code and party_code:
                        party_mapping[code] = party_code
        except Exception as e:
            print(f"Error fetching party names for summary: {e}")

        # Create Summary.xlsx workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb_summary = Workbook()
        ws_summary = wb_summary.active
        ws_summary.title = "Summary"
        
        row_idx = 1
        for rec in summary_records:
            prefix = rec['prefix']
            invoice_range = rec['invoice_range_str']
            
            # Lookup party name
            party_name = party_mapping.get(prefix, f"{prefix}-UNKNOWN")
            
            ws_summary.cell(row=row_idx, column=1, value=party_name).font = Font(name="Calibri", size=11, bold=True)
            ws_summary.cell(row=row_idx+1, column=1, value=invoice_range)
            
            row_idx += 3 # leave one empty cell
            
        ws_summary.column_dimensions['A'].width = 35
        wb_summary.save(os.path.join(temp_work_dir, "Summary.xlsx"))
        total_files_count += 1
        
        log_summary.append({
            'operation': 'Summary File Created',
            'value': 'Summary.xlsx',
            'status': 'Success'
        })
            
        # Create output ZIP archive from temp_work_dir (preserving folder structure)
        temp_zip_path = os.path.join('temp', 'invoice_output.zip')
        if os.path.exists(temp_zip_path):
            os.remove(temp_zip_path)
            
        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for root, dirs, files in os.walk(temp_work_dir):
                for file in files:
                    full_p = os.path.join(root, file)
                    rel_p = os.path.relpath(full_p, temp_work_dir)
                    z.write(full_p, rel_p)
                    
        shutil.rmtree(temp_work_dir, ignore_errors=True)
        
        return jsonify({
            'message': 'Invoice Arrange workflow completed successfully!',
            'files_count': total_files_count,
            'zip_filename': zip_display_name,
            'log': log_summary
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed processing Invoice Arrange: {str(e)}'}), 500

@app.route('/api/download-invoice-zip', methods=['GET'])
def download_invoice_zip():
    custom_filename = request.args.get('filename', 'Arranged_Invoices.zip')
    temp_path = os.path.join('temp', 'invoice_output.zip')
    
    if not os.path.exists(temp_path):
        return jsonify({'error': 'Arranged ZIP archive not found. Please upload and process again.'}), 400
        
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=custom_filename,
        mimetype='application/zip'
    )

# Hardcoded Apps Script URL
APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbzdq9Sps5C6QCVy5GGbeRpfM7f_3j_d7B9mL4cCBUFqb1avgn4vVUGW6F3kZMj5gWI0eg/exec"

def get_apps_script_url():
    config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    if not os.path.exists(config_path):
        return APPS_SCRIPT_URL
    import json
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            url = data.get('apps_script_url', '').strip()
            return url if url else APPS_SCRIPT_URL
    except Exception:
        return APPS_SCRIPT_URL

def save_apps_script_url(url):
    config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    import json
    config_data = {}
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
        except Exception:
            pass
    config_data['apps_script_url'] = url.strip()
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config_data, f, indent=4)

@app.route('/api/parties/config', methods=['GET', 'POST'])
def parties_config():
    if request.method == 'POST':
        data = request.get_json() or {}
        url = data.get('apps_script_url', '').strip()
        save_apps_script_url(url)
        return jsonify({'status': 'success', 'apps_script_url': get_apps_script_url()})
    else:
        return jsonify({'status': 'success', 'apps_script_url': get_apps_script_url()})

@app.route('/api/parties', methods=['GET'])
def get_parties():
    url = get_apps_script_url()
    import requests
    try:
        r = requests.get(url, allow_redirects=True, timeout=15)
        try:
            return jsonify(r.json())
        except Exception:
            preview = r.text[:250]
            return jsonify({
                'error': f'Google Apps Script returned non-JSON output. Please verify that the Apps Script is deployed as a Web App with access set to "Anyone" and authorized successfully. Response preview: {preview}'
            }), 400
    except Exception as e:
        return jsonify({'error': f'Failed connecting to Google Apps Script: {str(e)}'}), 500

@app.route('/api/parties/add', methods=['POST'])
def add_party():
    url = get_apps_script_url()
    data = request.get_json() or {}
    code = data.get('code', '').strip()
    partyCode = data.get('partyCode', '').strip()
    
    if not code or not partyCode:
        return jsonify({'error': 'Code and Party Code are required.'}), 400
        
    import requests
    try:
        payload = {
            'action': 'add',
            'code': code,
            'partyCode': partyCode
        }
        r = requests.post(url, json=payload, allow_redirects=True, timeout=15)
        return jsonify(r.json())
    except Exception as e:
        return jsonify({'error': f'Failed adding party: {str(e)}'}), 500

@app.route('/api/parties/update', methods=['POST'])
def update_party():
    url = get_apps_script_url()
    data = request.get_json() or {}
    rowIndex = data.get('rowIndex')
    code = data.get('code', '').strip()
    partyCode = data.get('partyCode', '').strip()
    
    if rowIndex is None or not code or not partyCode:
        return jsonify({'error': 'rowIndex, Code, and Party Code are required.'}), 400
        
    import requests
    try:
        payload = {
            'action': 'update',
            'rowIndex': rowIndex,
            'code': code,
            'partyCode': partyCode
        }
        r = requests.post(url, json=payload, allow_redirects=True, timeout=15)
        return jsonify(r.json())
    except Exception as e:
        return jsonify({'error': f'Failed updating party: {str(e)}'}), 500

@app.route('/api/parties/delete', methods=['POST'])
def delete_party():
    url = get_apps_script_url()
    data = request.get_json() or {}
    rowIndex = data.get('rowIndex')
    
    if rowIndex is None:
        return jsonify({'error': 'rowIndex is required.'}), 400
        
    import requests
    try:
        payload = {
            'action': 'delete',
            'rowIndex': rowIndex
        }
        r = requests.post(url, json=payload, allow_redirects=True, timeout=15)
        return jsonify(r.json())
    except Exception as e:
        return jsonify({'error': f'Failed deleting party: {str(e)}'}), 500

def process_flipkart_error(details_bytes, details_filename, data_bytes, data_filename, from_date_str, to_date_str):
    import datetime
    import shutil
    import tempfile
    
    from_date = None
    to_date = None
    if from_date_str and to_date_str:
        from_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
        to_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date()
        
    wb_details = load_file_to_openpyxl_workbook(details_bytes, details_filename)
    ws_details = wb_details.active
    
    df_data = load_any_sheet_to_dataframe(data_bytes, data_filename, header=None)
    data_lookup = {}
    if df_data is not None and not df_data.empty:
        for idx, row in df_data.iterrows():
            if len(row) > 4:
                key = str(row.iloc[4]).strip()
                val = str(row.iloc[2]).strip()
                if key:
                    data_lookup[key] = val
                    
    for r in range(ws_details.max_row, 1, -1):
        date_val = ws_details.cell(row=r, column=23).value
        delete_row = False
        
        if from_date and to_date and date_val:
            try:
                row_date = None
                if isinstance(date_val, datetime.datetime):
                    row_date = date_val.date()
                else:
                    date_str = str(date_val).strip().split(' ')[0]
                    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
                        try:
                            row_date = datetime.datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            pass
                
                if row_date and from_date <= row_date <= to_date:
                    delete_row = True
            except Exception:
                pass
                
        if not delete_row:
            v_val = ws_details.cell(row=r, column=22).value
            if v_val is not None:
                v_clean = str(v_val).strip()
                if v_clean == "0" or v_clean == "Price Dispute : 0":
                    delete_row = True
                    
        if delete_row:
            ws_details.delete_rows(r, 1)
            continue
            
        b_val = ws_details.cell(row=r, column=2).value
        if b_val:
            b_clean = str(b_val).strip()
            if b_clean in data_lookup:
                ws_details.cell(row=r, column=23, value=data_lookup[b_clean])
                
    party_data = {}
    for r in range(2, ws_details.max_row + 1):
        party_name = ws_details.cell(row=r, column=4).value
        if party_name:
            party_name = str(party_name).strip()
            if party_name.lower() == "warehouse name":
                continue
            if party_name not in party_data:
                party_data[party_name] = []
            row_vals = [ws_details.cell(row=r, column=c).value for c in range(1, ws_details.max_column + 1)]
            party_data[party_name].append(row_vals)
            
    temp_work_dir = tempfile.mkdtemp(dir='temp')
    party_files_info = []
    
    wb_master = Workbook()
    wb_master.remove(wb_master.active)
    
    from openpyxl.styles import Alignment, Font
    
    for party, rows in party_data.items():
        wb_party = Workbook()
        ws_party = wb_party.active
        
        sheet_title = f"{party}-price dispute"
        safe_sheet_title = sheet_title[:31]
        for char in ['\\', '/', '*', '?', ':', '[', ']']:
            safe_sheet_title = safe_sheet_title.replace(char, '')
            
        ws_party.title = safe_sheet_title
        ws_master_party = wb_master.create_sheet(title=safe_sheet_title)
        
        for ws in [ws_party, ws_master_party]:
            ws.merge_cells('A1:L1')
            cell = ws.cell(row=1, column=1, value=sheet_title)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=14)
            
        current_row = 2
        for r_data in rows:
            val_a = r_data[1] if len(r_data) > 1 else ""
            val_b = r_data[2] if len(r_data) > 2 else ""
            val_c = r_data[3] if len(r_data) > 3 else ""
            val_d = r_data[6] if len(r_data) > 6 else ""
            val_e = r_data[7] if len(r_data) > 7 else ""
            val_f = r_data[8] if len(r_data) > 8 else ""
            val_g = r_data[11] if len(r_data) > 11 else ""
            val_h = r_data[12] if len(r_data) > 12 else ""
            val_i = r_data[21] if len(r_data) > 21 else ""
            val_j = r_data[22] if len(r_data) > 22 else ""
            
            val_k = ""
            if val_i is not None and str(val_i).startswith("Price Dispute :"):
                try:
                    num_str = str(val_i).split(":")[1].strip()
                    dispute_num = float(num_str)
                    base_amount = float(str(val_h).strip()) if val_h else 0.0
                    val_k = base_amount - dispute_num
                except Exception:
                    val_k = "Error"
            
            val_l = "this amount not coorect as account central price this is approx price that currently live in account central"
            
            mapped_row = [val_a, val_b, val_c, val_d, val_e, val_f, val_g, val_h, val_i, val_j, val_k, val_l]
            ws_party.append(mapped_row)
            ws_master_party.append(mapped_row)
            
        from openpyxl.utils import get_column_letter
        for ws in [ws_party, ws_master_party]:
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for r_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
                
        party_filename = f"{sheet_title}.xlsx"
        party_filepath = os.path.join(temp_work_dir, party_filename)
        wb_party.save(party_filepath)
        party_files_info.append((party_filename, party_filepath))
        
    if not wb_master.sheetnames:
        wb_master.create_sheet(title="No Errors")
    master_filepath = os.path.join(temp_work_dir, "FLIPKART price dispute.xlsx")
    wb_master.save(master_filepath)
    
    for r in range(2, ws_details.max_row + 1):
        ws_details.cell(row=r, column=23, value="")
        
    cleaned_details_path = os.path.join(temp_work_dir, f"CLEANED_{details_filename}")
    wb_details.save(cleaned_details_path)
    
    import zipfile
    zip_path = os.path.join('temp', 'Flipkart_Error_Output.zip')
    if os.path.exists(zip_path):
        try: os.remove(zip_path)
        except: pass
        
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(cleaned_details_path, os.path.basename(cleaned_details_path))
        zf.write(master_filepath, os.path.basename(master_filepath))
        for p_name, p_path in party_files_info:
            zf.write(p_path, p_name)
            
    shutil.rmtree(temp_work_dir, ignore_errors=True)
    return zip_path

@app.route('/api/flipkart-error', methods=['POST'])
def flipkart_error_api():
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
        
    files = request.files.getlist('files[]')
    if len(files) != 2:
        return jsonify({'error': 'Exactly 2 files (Details and Data) are required.'}), 400
        
    from_date_str = request.form.get('fromDate')
    to_date_str = request.form.get('toDate')
    
    details_file = files[0]
    data_file = files[1]
    
    try:
        details_bytes = details_file.read()
        data_bytes = data_file.read()
        
        zip_path = process_flipkart_error(
            details_bytes, details_file.filename,
            data_bytes, data_file.filename,
            from_date_str, to_date_str
        )
        
        return jsonify({
            'message': 'Flipkart Error processed successfully!',
            'zip_path': zip_path
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed processing Flipkart Error: {str(e)}'}), 500

@app.route('/api/download-error-zip', methods=['GET'])
def download_error_zip():
    temp_path = os.path.join('temp', 'Flipkart_Error_Output.zip')
    if not os.path.exists(temp_path):
        return jsonify({'error': 'ZIP file not found.'}), 400
        
    return send_file(
        temp_path,
        as_attachment=True,
        download_name='Flipkart_Error_Output.zip',
        mimetype='application/zip'
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
