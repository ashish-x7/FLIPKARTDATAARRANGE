import os
import pandas as pd
from openpyxl import Workbook

def create_mock_excel(filename, start_id, num_rows):
    data = []
    statuses = ["DELIVERED", "RETURNED", "CANCELLED", "RETURN_REQUESTED", "READY_TO_SHIP"]
    
    for i in range(num_rows):
        item_numeric_id = start_id + i
        order_item_id = f"OI:{item_numeric_id}"
        order_id = f"OD{item_numeric_id}"
        
        # Double quotes around sku and product_title to mimic Flipkart export format
        sku = f'"""SKU:NK-828-Navy_{"M" if i%2==0 else "L"}"""'
        product_title = f'"""Vivatra Women Kurta Set {i}"""'
        
        row = {
            "order_item_id": order_item_id,
            "order_id": order_id,
            "fulfillment_source": "Seller",
            "fulfillment_type": "NON_FBF",
            "order_date": f"2026-06-{10 + i%15} 12:30:15",
            "order_approval_date": f"2026-06-{10 + i%15} 12:35:45",
            "order_item_status": statuses[i % len(statuses)],
            "sku": sku,
            "fsn": f"ETHHZ3ZSHSKG7{100+i}P9",
            "product_title": product_title
        }
        data.append(row)
        
    df = pd.DataFrame(data)
    
    os.makedirs('test_files', exist_ok=True)
    filepath = os.path.join('test_files', filename)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Orders', index=False)
        
    print(f"Created mock file: {filepath} with {num_rows} orders.")

def create_rename_test_files():
    os.makedirs('test_files', exist_ok=True)
    
    # 1. Create Mapping file (arrange_mapping.xlsx)
    # Sheet named 'ARRANGE'
    wb_map = Workbook()
    ws_map = wb_map.active
    ws_map.title = "ARRANGE"
    
    # Write headerless mapping
    # Column A (Code), Column B (Brand Name)
    ws_map.cell(row=1, column=1, value="CG1")
    ws_map.cell(row=1, column=2, value="NK-828-Navy")
    
    ws_map.cell(row=2, column=1, value="CG2")
    ws_map.cell(row=2, column=2, value="NK-933-Wine")
    
    ws_map.cell(row=3, column=1, value="CG3")
    ws_map.cell(row=3, column=2, value="NK-834-Off-White")
    
    map_path = os.path.join('test_files', 'arrange_mapping.xlsx')
    wb_map.save(map_path)
    print(f"Created mock mapping: {map_path}")
    
    # 2. Create File for Option A (P2 = NK-828-Navy)
    wb_a = Workbook()
    ws_a = wb_a.active
    # P2 cell is row 2, column 16 (P)
    ws_a.cell(row=2, column=16, value="NK-828-Navy")
    
    path_a = os.path.join('test_files', 'myntra_invoice_A.xlsx')
    wb_a.save(path_a)
    print(f"Created Option A test file: {path_a}")
    
    # 3. Create File for Option B (Column G has non-CGJ1 values)
    wb_b = Workbook()
    ws_b = wb_b.active
    # Column G is column 7
    # Row 1 is header
    ws_b.cell(row=1, column=7, value="Some_Column_G")
    ws_b.cell(row=2, column=7, value="CGJ1-SKIP-ME") # Should skip
    ws_b.cell(row=3, column=7, value="FLIPKART-828-Navy") # Split by '-' -> FLIPKART -> Right 3 -> ART
    
    path_b = os.path.join('test_files', 'flipkart_invoice_B.xlsx')
    wb_b.save(path_b)
    print(f"Created Option B test file: {path_b}")

if __name__ == "__main__":
    create_mock_excel("flipkart_orders_batch1.xlsx", 337786218019885100, 10)
    create_mock_excel("flipkart_orders_batch2.xlsx", 337786218019885200, 5)
    create_rename_test_files()
