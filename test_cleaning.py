import os
import pandas as pd
import io
from openpyxl import load_workbook
from app import (
    clean_order_item_id, 
    clean_quotes, 
    clean_sku, 
    find_col_key,
    load_saved_mapping,
    find_rename_code_option_a_from_bytes,
    find_rename_code_option_b_from_bytes,
    clean_string_for_matching
)

def run_tests():
    print("Running cleaning and merging tests...")
    
    # 1. Test individual cleaning functions
    assert clean_order_item_id("OI:337786218019885100") == "337786218019885100"
    assert clean_order_item_id("oi:12345") == "12345"
    assert clean_order_item_id("  OI:98765  ") == "98765"
    assert clean_order_item_id("88888") == "88888"
    assert clean_order_item_id(None) == ""
    
    assert clean_sku('"""SKU:NK-828-Navy_M"""') == "NK-828-Navy_M"
    assert clean_sku('""SKU:NK-828-Navy_M""') == "NK-828-Navy_M"
    assert clean_sku('"SKU:NK-828-Navy_M"') == "NK-828-Navy_M"
    assert clean_sku('"NK-828-Navy_M"') == "NK-828-Navy_M"
    assert clean_sku(None) == ""
    
    assert clean_quotes('"""Vivatra Women Kurta"""') == "Vivatra Women Kurta"
    assert clean_quotes(None) == ""
    
    # 2. Check if mock files exist
    batch1_path = os.path.join("test_files", "flipkart_orders_batch1.xlsx")
    batch2_path = os.path.join("test_files", "flipkart_orders_batch2.xlsx")
    
    if not os.path.exists(batch1_path) or not os.path.exists(batch2_path):
        print("Mock files not found! Make sure generate_mock_data.py has finished running.")
        return
        
    print("Found mock files. Simulating upload & process logic...")
    
    # Load and clean batch 1
    xls1 = pd.ExcelFile(batch1_path)
    df1 = xls1.parse("Orders", dtype=str)
    
    col_id1 = find_col_key(df1, ['order_item_id'], 0)
    col_sku1 = find_col_key(df1, ['sku'], 7)
    col_title1 = find_col_key(df1, ['product_title'], 9)
    
    df1[col_id1] = df1[col_id1].apply(clean_order_item_id)
    df1[col_sku1] = df1[col_sku1].apply(clean_sku)
    df1[col_title1] = df1[col_title1].apply(clean_quotes)
    
    # Load and clean batch 2
    xls2 = pd.ExcelFile(batch2_path)
    df2 = xls2.parse("Orders", dtype=str)
    
    col_id2 = find_col_key(df2, ['order_item_id'], 0)
    col_sku2 = find_col_key(df2, ['sku'], 7)
    col_title2 = find_col_key(df2, ['product_title'], 9)
    
    df2[col_id2] = df2[col_id2].apply(clean_order_item_id)
    df2[col_sku2] = df2[col_sku2].apply(clean_sku)
    df2[col_title2] = df2[col_title2].apply(clean_quotes)
    
    # Merge dataframes
    combined = pd.concat([df1, df2], ignore_index=True).fillna('')
    assert len(combined) == 15, f"Expected 15 merged rows, got {len(combined)}"
    
    # Save test output
    test_out_path = os.path.join("test_files", "test_output_merged.xlsx")
    
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Determine columns
    order_item_col_idx = list(combined.columns).index(col_id1) + 1
    sku_col_idx = list(combined.columns).index(col_sku1) + 1
    title_col_idx = list(combined.columns).index(col_title1) + 1
    
    for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if c_idx == order_item_col_idx and r_idx > 1:
                cell.value = str(val).strip()
                cell.data_type = 's'  # Enforce Text
                cell.number_format = '@'
            else:
                cell.value = val
                
    wb.save(test_out_path)
    print(f"Saved test merged file to {test_out_path}")
    
    # Verify using raw openpyxl
    wb_read = load_workbook(test_out_path)
    ws_read = wb_read["Orders"]
    
    cell_a2 = ws_read.cell(row=2, column=1)
    assert cell_a2.data_type == 's'
    assert cell_a2.value == "337786218019885100"
    
    cell_h2 = ws_read.cell(row=2, column=sku_col_idx)
    assert cell_h2.value == "NK-828-Navy_M"
    
    cell_j2 = ws_read.cell(row=2, column=title_col_idx)
    assert cell_j2.value == "Vivatra Women Kurta Set 0"
    
    # ----------------------------------------------------
    # RENAME FILES TESTS
    # ----------------------------------------------------
    print("\nRunning File Renaming Logic tests...")
    
    # Test clean_string_for_matching
    assert clean_string_for_matching("NK-828-Navy") == "NK828NAVY"
    assert clean_string_for_matching("NK_828.Navy ") == "NK828NAVY"
    
    # Set up mapping path in temp directory for test
    os.makedirs('temp', exist_ok=True)
    import shutil
    shutil.copyfile(
        os.path.join('test_files', 'arrange_mapping.xlsx'),
        os.path.join('temp', 'arrange_mapping.xlsx')
    )
    
    # Test load_saved_mapping
    mapping_df = load_saved_mapping()
    assert mapping_df is not None, "Failed to load arrange mapping"
    assert len(mapping_df) == 3, f"Expected 3 mapping rules, got {len(mapping_df)}"
    print("Successfully verified load_saved_mapping!")
    
    # Test Option A Renamer (using myntra_invoice_A.xlsx where cell P2 = NK-828-Navy)
    with open(os.path.join('test_files', 'myntra_invoice_A.xlsx'), 'rb') as f:
        file_bytes_a = f.read()
    rename_code_a = find_rename_code_option_a_from_bytes(file_bytes_a, 'myntra_invoice_A.xlsx', mapping_df)
    print(f"Option A Extracted rename code: {rename_code_a}")
    assert rename_code_a == "CG1", f"Expected CG1, got {rename_code_a}"
    
    # Test Option B Renamer (using flipkart_invoice_B.xlsx where Col G has CGJ1-SKIP-ME and FLIPKART-828-Navy)
    with open(os.path.join('test_files', 'flipkart_invoice_B.xlsx'), 'rb') as f:
        file_bytes_b = f.read()
    rename_code_b = find_rename_code_option_b_from_bytes(file_bytes_b, 'flipkart_invoice_B.xlsx')
    print(f"Option B Extracted rename code: {rename_code_b}")
    assert rename_code_b == "ART", f"Expected ART, got {rename_code_b}"
    
    # ----------------------------------------------------
    # SPLIT FILES TESTS
    # ----------------------------------------------------
    print("\nRunning Split Files Logic tests...")
    from app import run_split_logic_from_bytes
    
    # Create a mock spreadsheet for splitting (Option 1 & 2 & 3: 2 header rows, filter starts on row 3)
    wb_split = Workbook()
    ws_split = wb_split.active
    ws_split.title = "Sheet1"
    
    # Headers
    ws_split.cell(row=1, column=4, value="HEADER_L1_COL_D")
    ws_split.cell(row=1, column=7, value="HEADER_L1_COL_G")
    ws_split.cell(row=2, column=4, value="Warehouse AJIO") # Column D
    ws_split.cell(row=2, column=7, value="Warehouse FLIPKART") # Column G
    
    # Row 3 (Data Row 1)
    ws_split.cell(row=3, column=4, value="MUMBAI-AJIO")
    ws_split.cell(row=3, column=7, value="DELHI-FLIPKART")
    
    # Row 4 (Data Row 2)
    ws_split.cell(row=4, column=4, value="BANGALORE-AJIO")
    ws_split.cell(row=4, column=7, value="DELHI-FLIPKART")
    
    # Row 5 (Data Row 3)
    ws_split.cell(row=5, column=4, value="MUMBAI-AJIO")
    ws_split.cell(row=5, column=7, value="KOLKATA-FLIPKART")
    
    split_io = io.BytesIO()
    wb_split.save(split_io)
    split_bytes = split_io.getvalue()
    
    # Test Option 1 (Simple) - Filter D (AJIO/FLIPKART)
    files_opt1_d = run_split_logic_from_bytes(split_bytes, "test_split.xlsx", "1")
    print(f"Option 1 D Split files count: {len(files_opt1_d)}")
    assert len(files_opt1_d) == 2, f"Expected 2 split files, got {len(files_opt1_d)}"
    # Mumbai and Bangalore
    opt1_d_names = [f[0] for f in files_opt1_d]
    assert any("MUMBAI-AJIO-FLIPKART" in n for n in opt1_d_names)
    assert any("BANGALORE-AJIO-FLIPKART" in n for n in opt1_d_names)

    # Test Option 2 (Details) - Filter D (AJIO)
    files_opt2 = run_split_logic_from_bytes(split_bytes, "test_split.xlsx", "2", "")
    print(f"Option 2 Split files count: {len(files_opt2)}")
    assert len(files_opt2) == 2, f"Expected 2 split files, got {len(files_opt2)}"
    opt2_names = [f[0] for f in files_opt2]
    assert any("MUMBAI-AJIO DETAILS SHEET FLIPKART" in n for n in opt2_names)
    assert any("BANGALORE-AJIO DETAILS SHEET FLIPKART" in n for n in opt2_names)

    # Test Option 3 (Summary) - Filter G (FLIPKART)
    files_opt3 = run_split_logic_from_bytes(split_bytes, "test_split.xlsx", "3", "")
    print(f"Option 3 Split files count: {len(files_opt3)}")
    assert len(files_opt3) == 2, f"Expected 2 split files, got {len(files_opt3)}"
    opt3_names = [f[0] for f in files_opt3]
    assert any("DELHI-FLIPKART SUMMARY SHEET FLIPKART" in n for n in opt3_names)
    assert any("KOLKATA-FLIPKART SUMMARY SHEET FLIPKART" in n for n in opt3_names)

    # Test Option 4 (Tax Split) - 1 header row, filter Column A (1)
    wb_tax = Workbook()
    ws_tax = wb_tax.active
    ws_tax.title = "Sheet1"
    # Header
    ws_tax.cell(row=1, column=1, value="GSTIN")
    ws_tax.cell(row=1, column=2, value="Amount")
    # Data
    ws_tax.cell(row=2, column=1, value="157-INDOPRIMO")
    ws_tax.cell(row=2, column=2, value="1000")
    ws_tax.cell(row=3, column=1, value="158-JACKJONES")
    ws_tax.cell(row=3, column=2, value="2000")
    ws_tax.cell(row=4, column=1, value="157-INDOPRIMO")
    ws_tax.cell(row=4, column=2, value="3000")
    
    tax_io = io.BytesIO()
    wb_tax.save(tax_io)
    tax_bytes = tax_io.getvalue()
    
    files_opt4 = run_split_logic_from_bytes(tax_bytes, "test_tax.xlsx", "4", "")
    print(f"Option 4 Split files count: {len(files_opt4)}")
    assert len(files_opt4) == 2, f"Expected 2 split files, got {len(files_opt4)}"
    opt4_names = [f[0] for f in files_opt4]
    # format: firstNum & "-Tax-" & whName & "-MYNTYRA"
    # firstNum of 157-INDOPRIMO is 157
    assert any("157-Tax-157-INDOPRIMO-MYNTYRA" in n for n in opt4_names)
    assert any("158-Tax-158-JACKJONES-MYNTYRA" in n for n in opt4_names)

    # ----------------------------------------------------
    # CREATE FOLDER API TESTS
    # ----------------------------------------------------
    print("\nRunning Create Folder API tests...")
    from app import app
    client = app.test_client()
    
    merged_data = io.BytesIO(b"dummy merged data")
    prefix1_data = io.BytesIO(b"dummy prefix 101 data")
    prefix2_data = io.BytesIO(b"dummy prefix 509 data")
    
    data = {
        'files[]': [
            (merged_data, 'Flipkart_Merged_Orders.xlsx'),
            (prefix1_data, '101-TaxReportData.xlsx'),
            (prefix2_data, '509-TaxReportData.xlsx')
        ]
    }
    
    resp = client.post('/api/create-folder', data=data, content_type='multipart/form-data')
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
    res_data = resp.get_json()
    print("API Response:", res_data)
    assert res_data['folders_count'] == 2
    assert res_data['zip_filename'] == '101-509.zip'
    assert len(res_data['log']) == 2
    
    log_101 = next(item for item in res_data['log'] if item['folder'] == '101')
    assert log_101['copied_merged'] == 'Flipkart_Merged_Orders.xlsx'
    assert '101-TaxReportData.xlsx' in log_101['moved_files']

    # ----------------------------------------------------
    # INVOICE ARRANGE API TESTS
    # ----------------------------------------------------
    print("\nRunning Invoice Arrange API tests...")
    
    wb_det = Workbook()
    ws_det = wb_det.active
    ws_det.title = "Sheet1"
    ws_det.cell(row=1, column=1, value="Sale Item Details")
    ws_det.cell(row=2, column=1, value="No.")
    ws_det.cell(row=2, column=2, value="Invoice No")
    ws_det.cell(row=3, column=1, value="1")
    ws_det.cell(row=3, column=2, value="FK27S101-1")
    
    det_io = io.BytesIO()
    wb_det.save(det_io)
    det_bytes = det_io.getvalue()
    
    wb_dt_test = Workbook()
    ws_dt_test = wb_dt_test.active
    ws_dt_test.title = "TaxReportData"
    
    headers_dt = ["Company Name", "Seller GST Num", "MP Name", "Reference Code", "Suborder No", "Order Type", "EE Invoice No", "MP Ref No", "Order Status", "Invoice Status", "Import Date", "Order Date"]
    ws_dt_test.append(headers_dt)
    
    ws_dt_test.append(["101-B", "GST1", "Flipkart", "OD33", "337949622155640100", "B2C", "FK27S101-1", "REF1", "Shipped", "Sold", "2026", "2026"])
    
    ws_dt_test.cell(row=3, column=1, value="101-B")
    ws_dt_test.cell(row=3, column=5, value="`337949622155640200")
    ws_dt_test.cell(row=3, column=7, value="FK27S101-2")
    ws_dt_test.cell(row=3, column=9, value="Shipped")
    ws_dt_test.cell(row=3, column=13, value="20/06/2026")
    ws_dt_test.cell(row=3, column=18, value="1")
    ws_dt_test.cell(row=3, column=40, value="Jammu & Kashmir") # Column AN (State)
    ws_dt_test.cell(row=3, column=42, value="18")      # Column AP (GST Rate)
    ws_dt_test.cell(row=3, column=48, value="100") # AV
    ws_dt_test.cell(row=3, column=50, value="80")  # AX
    
    ws_dt_test.cell(row=4, column=1, value="101-B")
    ws_dt_test.cell(row=4, column=5, value="337949622155640200")
    ws_dt_test.cell(row=4, column=7, value="FK27S101-2")
    ws_dt_test.cell(row=4, column=9, value="Shipped")
    ws_dt_test.cell(row=4, column=13, value="20/06/2026")
    ws_dt_test.cell(row=4, column=18, value="1")
    ws_dt_test.cell(row=4, column=40, value="Gujarat") # Column AN (State)
    ws_dt_test.cell(row=4, column=42, value="18")      # Column AP (GST Rate)
    ws_dt_test.cell(row=4, column=48, value="100") # AV
    ws_dt_test.cell(row=4, column=50, value="80")  # AX
    
    dt_test_io = io.BytesIO()
    wb_dt_test.save(dt_test_io)
    dt_test_bytes = dt_test_io.getvalue()
    
    wb_od_test = Workbook()
    ws_od_test = wb_od_test.active
    ws_od_test.title = "Orders"
    ws_od_test.append(["order_item_id", "colB", "colC", "colD", "colE", "colF", "colG", "sku", "fsn", "product_title"])
    ws_od_test.append(["337949622155640200", "", "", "", "", "", "", "SKU1", "FSN1", "TITLE1"])
    
    od_test_io = io.BytesIO()
    wb_od_test.save(od_test_io)
    od_test_bytes = od_test_io.getvalue()
    
    det_io.seek(0)
    dt_test_io.seek(0)
    od_test_io.seek(0)
    
    invoice_data = {
        'OD': (od_test_io, 'Flipkart_Merged_Orders.xlsx'),
        'DT': (dt_test_io, '101-TaxReportData.xlsx'),
        'Details': (det_io, '101-Bharvita-FLIPKART.xlsx')
    }
    
    resp_inv = client.post('/api/invoice-arrange', data=invoice_data, content_type='multipart/form-data')
    assert resp_inv.status_code == 200, f"Expected 200, got {resp_inv.status_code}"
    res_inv_data = resp_inv.get_json()
    print("Invoice Arrange API Response:", res_inv_data)
    
    assert res_inv_data['zip_filename'] == "101-(FK27S101-2)-arranged.zip"
    assert any(log['operation'] == 'Matched Rows Deleted' and log['value'] == '1' for log in res_inv_data['log'])
    assert any(log['operation'] == 'Matched Suborders Highlighted' and log['value'] == '1' for log in res_inv_data['log'])
    assert any(log['operation'] == 'Output PR File Created' and 'PR.xlsx' in log['value'] for log in res_inv_data['log'])

    # ----------------------------------------------------
    # INVOICE ARRANGE BATCH ZIP TESTS
    # ----------------------------------------------------
    print("\nRunning Invoice Arrange Batch ZIP tests...")
    import zipfile
    
    zip_io = io.BytesIO()
    with zipfile.ZipFile(zip_io, 'w') as z:
        # Folder 101 files
        z.writestr("101/Flipkart_Merged_Orders.xlsx", od_test_bytes)
        z.writestr("101/101-TaxReportData.xlsx", dt_test_bytes)
        z.writestr("101/101-Bharvita-FLIPKART.xlsx", det_bytes)
        
        # Folder 509 files
        z.writestr("509/Flipkart_Merged_Orders.xlsx", od_test_bytes)
        z.writestr("509/509-TaxReportData.xlsx", dt_test_bytes)
        z.writestr("509/509-Bharvita-FLIPKART.xlsx", det_bytes)
        
    zip_io.seek(0)
    
    batch_data = {
        'zipfile': (zip_io, 'Batch_Invoices.zip')
    }
    
    resp_batch = client.post('/api/invoice-arrange', data=batch_data, content_type='multipart/form-data')
    assert resp_batch.status_code == 200, f"Expected 200, got {resp_batch.status_code}"
    res_batch_data = resp_batch.get_json()
    print("Batch Invoice Arrange API Response:", res_batch_data)
    assert res_batch_data['zip_filename'] == "101-509-arranged.zip"
    assert len(res_batch_data['log']) == 5 # 2 folders * 2 logs each + 1 summary log = 5

    # ----------------------------------------------------
    # PARTY DATA API TESTS (MOCK GOOGLE APPS SCRIPT)
    # ----------------------------------------------------
    print("\nRunning Party Data Proxy API tests...")
    
    original_config_content = None
    config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                original_config_content = f.read()
        except Exception:
            pass

    try:
        resp_config = client.get('/api/parties/config')
        assert resp_config.status_code == 200
        config_json = resp_config.get_json()
        assert 'apps_script_url' in config_json
        
        from app import APPS_SCRIPT_URL
        resp_save_config = client.post('/api/parties/config', json={'apps_script_url': 'https://mock.script.google.com/exec'})
        assert resp_save_config.status_code == 200
        assert resp_save_config.get_json()['status'] == 'success'
        
        resp_config2 = client.get('/api/parties/config')
        assert resp_config2.get_json()['apps_script_url'] == 'https://mock.script.google.com/exec'
        
        class MockResponse:
            def __init__(self, json_data, status_code=200):
                self.json_data = json_data
                self.status_code = status_code
                
            def json(self):
                return self.json_data
                
        mock_db = [
            {'CODE': '101', 'PARTY CODE': '101-BHARVITA', 'row_index': 2},
            {'CODE': '509', 'PARTY CODE': '509-VIVATRA', 'row_index': 3}
        ]
        
        def mock_get(url, *args, **kwargs):
            return MockResponse(mock_db)
            
        def mock_post(url, json=None, *args, **kwargs):
            action = json.get('action')
            if action == 'add':
                mock_db.append({'CODE': json['code'], 'PARTY CODE': json['partyCode'], 'row_index': 4})
                return MockResponse({'status': 'success', 'message': 'Party added successfully'})
            elif action == 'update':
                idx = int(json['rowIndex']) - 2
                mock_db[idx]['CODE'] = json['code']
                mock_db[idx]['PARTY CODE'] = json['partyCode']
                return MockResponse({'status': 'success', 'message': 'Party updated successfully'})
            elif action == 'delete':
                idx = int(json['rowIndex']) - 2
                mock_db.pop(idx)
                return MockResponse({'status': 'success', 'message': 'Party deleted successfully'})
            return MockResponse({'status': 'error', 'message': 'invalid action'})
            
        import requests
        original_get = requests.get
        original_post = requests.post
        
        requests.get = mock_get
        requests.post = mock_post
        
        try:
            resp_list = client.get('/api/parties')
            assert resp_list.status_code == 200
            parties_list = resp_list.get_json()
            assert len(parties_list) == 2
            assert parties_list[0]['CODE'] == '101'
            
            resp_add = client.post('/api/parties/add', json={'code': '102', 'partyCode': '102-MOCK'})
            assert resp_add.status_code == 200
            assert resp_add.get_json()['status'] == 'success'
            assert len(mock_db) == 3
            
            resp_up = client.post('/api/parties/update', json={'rowIndex': 2, 'code': '101-NEW', 'partyCode': '101-BHARVITA-NEW'})
            assert resp_up.status_code == 200
            assert resp_up.get_json()['status'] == 'success'
            assert mock_db[0]['CODE'] == '101-NEW'
            
            resp_del = client.post('/api/parties/delete', json={'rowIndex': 4})
            assert resp_del.status_code == 200
            assert resp_del.get_json()['status'] == 'success'
            assert len(mock_db) == 2
            
        finally:
            requests.get = original_get
            requests.post = original_post

    finally:
        try:
            if original_config_content is not None:
                with open(config_path, 'w', encoding='utf-8') as f:
                    f.write(original_config_content)
            elif os.path.exists(config_path):
                os.remove(config_path)
        except Exception:
            pass

    print("\n--- ALL TESTS (MERGING + RENAMING + SPLITTING + GROUPING + INVOICE_ARRANGE + BATCH_ZIP + PARTY_DATA) PASSED SUCCESSFULLY! ---")

if __name__ == "__main__":
    run_tests()
